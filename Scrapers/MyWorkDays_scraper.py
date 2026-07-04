import json
import time
import requests
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Config
TENANT = "ocbc"
SITE = "External"
BASE = f"https://{TENANT}.wd102.myworkdayjobs.com"
API_URL = f"{BASE}/wday/cxs/{TENANT}/{SITE}/jobs"
JOB_PAGE_BASE = f"{BASE}/en-US/{SITE}/job"

PAGE_SIZE = 20          # Workday's default page size for this endpoint
REQUEST_DELAY = 0.5     # be polite between API calls
DETAIL_DELAY = 0.8      # delay between job-detail page loads
OUTPUT_FILE = "ocbc_jobs.xlsx"


# Phase 1: Get all job listings via the Workday JSON API
def fetch_all_listings():
    """
    Paginate through the Workday CXS jobs API and collect basic info
    for every job: title, location, posted date, and the path used to
    build the detail-page link.
    """
    session = requests.Session()
    session.headers.update({
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    })

    listings = []
    offset = 0
    total = None

    while True:
        payload = {
            "appliedFacets": {},
            "limit": PAGE_SIZE,
            "offset": offset,
            "searchText": ""
        }
        resp = session.post(API_URL, json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        if total is None:
            total = data.get("total", 0)
            print(f"Total jobs found: {total}")

        job_postings = data.get("jobPostings", [])
        if not job_postings:
            break

        for job in job_postings:
            title = job.get("title", "")
            location = job.get("locationsText", "") or job.get("location", "")
            posted_date = job.get("postedOn", "")
            external_path = job.get("externalPath", "")  # e.g. /job/Singapore/Title_REQID
            job_id = job.get("bulletFields", [None])[0] if job.get("bulletFields") else ""

            full_link = f"{BASE}/en-US/{SITE}{external_path}" if external_path else ""

            listings.append({
                "title": title,
                "location": location,
                "posted_date": posted_date,
                "job_id": job_id,
                "link": full_link,
            })

        offset += PAGE_SIZE
        print(f"Fetched {len(listings)} / {total}")

        if offset >= total:
            break

        time.sleep(REQUEST_DELAY)

    return listings


# Phase 2: Visit each job page and scrape full details
def scrape_job_details(listings):
    """
    Opens each job's detail page with Playwright and extracts the full
    job description plus any structured fields Workday exposes
    (job ID, time type, posted date confirmation, etc).
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for i, job in enumerate(listings, 1):
            if not job["link"]:
                job["description"] = ""
                continue

            try:
                page.goto(job["link"], wait_until="domcontentloaded", timeout=30000)
                page.wait_for_selector(
                    '[data-automation-id="jobPostingDescription"]',
                    timeout=15000
                )

                description = page.locator(
                    '[data-automation-id="jobPostingDescription"]'
                ).inner_text()

                # Optional structured fields commonly present on Workday job pages
                def safe_text(selector):
                    try:
                        return page.locator(selector).first.inner_text()
                    except Exception:
                        return ""

                job["description"] = description
                job["job_req_id"] = safe_text('[data-automation-id="job-posting-id"]')
                job["time_type"] = safe_text('[data-automation-id="time"]')

                print(f"[{i}/{len(listings)}] Scraped: {job['title']}")

            except Exception as e:
                print(f"[{i}/{len(listings)}] FAILED: {job['title']} -> {e}")
                job["description"] = ""

            time.sleep(DETAIL_DELAY)

        browser.close()

    return listings


# Phase 3: Write to Excel
def write_excel(listings, filename=OUTPUT_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCBC Jobs"

    headers = ["Title", "Location", "Posted Date", "Job Req ID", "Time Type", "Link", "Description"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for job in listings:
        ws.append([
            job.get("title", ""),
            job.get("location", ""),
            job.get("posted_date", ""),
            job.get("job_req_id", job.get("job_id", "")),
            job.get("time_type", ""),
            job.get("link", ""),
            job.get("description", ""),
        ])

    widths = [40, 25, 15, 18, 12, 50, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(filename)
    print(f"Saved {len(listings)} jobs to {filename}")


# --- Main ---------------------------------------------------------------
if __name__ == "__main__":
    print("Phase 1: Fetching job listings from Workday API...")
    listings = fetch_all_listings()

    # Save a raw checkpoint in case Phase 2 fails partway through
    with open("listings_checkpoint.json", "w") as f:
        json.dump(listings, f, indent=2)

    print(f"\nPhase 2: Scraping {len(listings)} job detail pages...")
    listings = scrape_job_details(listings)

    print("\nPhase 3: Writing Excel file...")
    write_excel(listings)